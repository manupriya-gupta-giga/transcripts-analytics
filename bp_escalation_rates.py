#!/usr/bin/env python3
"""Escalation rate by intent around each billing-period (BP) day.

BP Day = the last day of a month. For each BP day and each intent, computes
the escalation rate (escalated / total chats, i.e. 1 - containment) for every
relative day T-7 .. T+9, where T is the BP day.

Day definition matches the ops dashboard ("Total Chats by Chat Date" sheet):
- Days are **US Eastern calendar days** (transcript createdAt converted to
  America/New_York). Raw downloads are foldered by UTC day, so each Eastern
  day D reads UTC sources D and D+1 and filters by Eastern date.
- Chats are **engaged only**: conversations where the customer sent at least
  one message (role == 1). Greeting-only chats are dropped. Disable with
  --keep-unengaged.

Verified 2026-08-17 against the dashboard sheet: Eastern + engaged, without
user exclusions, matches its daily totals within ~0.2% (several days exact).

Output: one sheet, rows = (BP Day, Intent), columns = T-7 .. T+9.

Sources, per UTC day: raw_transcripts/<date>/ when present (escalation =
create_ticket tool call, exact), else ticket_metadata/<date>.jsonl from
pull_ticket_metadata.py (escalation via cf flag or validated summary proxy).
Customer Support Chat Agent only. Metadata records carry neither customer
IDs nor messages, so --exclude-users and the engagement filter cannot apply
to them; affected days are reported.

Example:
  python3 bp_escalation_rates.py --bp-days 2026-06-30 2026-07-31 \
      --exclude-users exclude_user_ids.txt
"""

import argparse
import json
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo

HERE = Path(__file__).resolve().parent
AGENT_NAME = "Customer Support Chat Agent"
AGENT_TEMPLATE_ID = "24bd649e-6a44-4efd-af97-bfaf4dadfb43"
REL_DAYS = list(range(-7, 10))  # T-7 .. T+9
SKIP_DERIVE_TAGS = {"GreetingOnly", "RequestHumanSupport"}
EASTERN = ZoneInfo("America/New_York")


def tag_to_intent():
    path = HERE / "tag_taxonomy.json"
    if not path.exists():
        return {}
    return {t["name"]: t["intentName"]
            for t in json.loads(path.read_text())["tags"]
            if t["name"] not in SKIP_DERIVE_TAGS}


def resolve_intent(intent, tags, taxonomy):
    if not intent:
        intent = next((taxonomy[tag] for tag in tags if tag in taxonomy), None)
    return intent or "(no intent)"


def customer_id(transcript):
    for m in transcript.get("messages") or []:
        store = m.get("variableStore") or {}
        cid = store.get("customer_id") or store.get("external_id")
        if cid:
            return cid
    return None


def eastern_date(ts):
    try:
        return datetime.fromisoformat(ts.replace("Z", "+00:00")) \
                       .astimezone(EASTERN).date()
    except (AttributeError, ValueError, TypeError):
        return None


def load_utc_day(raw_dir: Path, meta_dir: Path, day: date, taxonomy):
    """Normalized records for one UTC day from the best available source.

    Each record: created (ET date), intent, escalated, engaged (bool, or
    None when the source has no messages), cid (or None), src.
    Returns None when neither source exists for the day.
    """
    records = []

    folder = raw_dir / day.isoformat()
    if folder.exists() and any(folder.glob("*.json")):
        for p in folder.glob("*.json"):
            try:
                t = json.loads(p.read_text(encoding="utf-8"))
            except (OSError, json.JSONDecodeError):
                continue
            if not (t.get("agentTemplateId") == AGENT_TEMPLATE_ID
                    or t.get("agentName") == AGENT_NAME):
                continue
            analysis = t.get("analysis") or {}
            msgs = t.get("messages") or []
            records.append({
                "created": eastern_date(t.get("createdAt")),
                "intent": resolve_intent(analysis.get("intent"),
                                         analysis.get("tags") or [], taxonomy),
                "escalated": any(
                    (tc.get("name") or (tc.get("function") or {}).get("name")) == "create_ticket"
                    for m in msgs for tc in m.get("toolCalls") or []),
                "engaged": any(m.get("role") == 1 for m in msgs),
                "cid": customer_id(t),
                "src": "raw",
            })
        return records

    meta = meta_dir / f"{day.isoformat()}.jsonl"
    if meta.exists():
        for line in meta.read_text().splitlines():
            try:
                r = json.loads(line)
            except json.JSONDecodeError:
                continue
            records.append({
                "created": eastern_date(r.get("createdAt")),
                "intent": resolve_intent(r.get("intent"), r.get("tags") or [], taxonomy),
                "escalated": bool(r.get("escalated")),
                "engaged": None,   # metadata has no messages
                "cid": None,       # metadata has no customer id
                "src": "meta",
            })
        return records
    return None


def day_stats(raw_dir: Path, meta_dir: Path, day: date, taxonomy,
              exclude_users=frozenset(), engaged_only=True, report=None,
              _cache={}):
    """{intent: [total, escalated]} for one US-Eastern day; None if no data.

    Reads UTC sources for `day` and `day + 1` and keeps records whose
    Eastern date is `day`. Exclusion and engagement filters apply to raw
    records; metadata records (no cid, no messages) pass through and are
    reported so the gap is visible.
    """
    stats = defaultdict(lambda: [0, 0])
    any_source = False
    for utc_day in (day, day + timedelta(days=1)):
        key = (str(raw_dir), str(meta_dir), utc_day)
        if key not in _cache:
            _cache[key] = load_utc_day(raw_dir, meta_dir, utc_day, taxonomy)
        records = _cache[key]
        if records is None:
            continue
        any_source = True
        for rec in records:
            if rec["created"] != day:
                continue
            if rec["src"] == "meta" and report is not None:
                if exclude_users:
                    report["unfilterable_days"].add(day.isoformat())
                if engaged_only:
                    report["unengaged_unknown_days"].add(day.isoformat())
            if exclude_users and rec["cid"] in exclude_users and rec["cid"]:
                if report is not None:
                    report["excluded"] += 1
                continue
            if engaged_only and rec["engaged"] is False:
                if report is not None:
                    report["unengaged_dropped"] += 1
                continue
            for k in (rec["intent"], "ALL"):
                stats[k][0] += 1
                stats[k][1] += int(rec["escalated"])
    return stats if any_source else None


def main():
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--bp-days", nargs="+", type=date.fromisoformat,
                        default=[date(2026, 6, 30), date(2026, 7, 31)],
                        help="BP days (last day of month), e.g. 2026-06-30")
    parser.add_argument("--raw-dir", type=Path, default=HERE / "raw_transcripts")
    parser.add_argument("--meta-dir", type=Path, default=HERE / "ticket_metadata")
    parser.add_argument("--min-chats", type=int, default=5,
                        help="Blank out cells with fewer chats than this")
    parser.add_argument("--top-intents", type=int, default=8,
                        help="Keep only the N highest-volume intents (plus ALL)")
    parser.add_argument("--exclude-users", type=Path, default=None,
                        help="File with one customer_id per line; their conversations "
                             "are dropped (raw-transcript days only — metadata "
                             "days have no customer id)")
    parser.add_argument("--keep-unengaged", action="store_true",
                        help="Keep chats where the customer never sent a message "
                             "(default drops them, matching the ops dashboard)")
    parser.add_argument("--out", type=Path, default=None)
    args = parser.parse_args()

    taxonomy = tag_to_intent()
    out = args.out or HERE / "results" / "escalation_rate_breakdown.xlsx"
    engaged_only = not args.keep_unengaged

    exclude_users = frozenset()
    if args.exclude_users:
        exclude_users = frozenset(
            line.strip() for line in args.exclude_users.read_text().splitlines()
            if line.strip())
        print(f"excluding {len(exclude_users)} user ids")
    report = {"excluded": 0, "unengaged_dropped": 0,
              "unfilterable_days": set(), "unengaged_unknown_days": set()}

    # per BP day: rel_day -> {intent: [total, escalated]}
    per_bp = {}
    intents = set()
    for bp in args.bp_days:
        per_bp[bp] = {}
        for rel in REL_DAYS:
            stats = day_stats(args.raw_dir, args.meta_dir, bp + timedelta(days=rel),
                              taxonomy, exclude_users, engaged_only, report)
            per_bp[bp][rel] = stats
            if stats:
                intents.update(stats.keys())

    if exclude_users:
        print(f"excluded {report['excluded']} conversations from raw-transcript days")
        if report["unfilterable_days"]:
            print("WARNING: no customer ids in metadata fallback — exclusion NOT "
                  "applied on:", ", ".join(sorted(report["unfilterable_days"])))
    if engaged_only:
        print(f"dropped {report['unengaged_dropped']} unengaged chats "
              "(no customer message)")
        if report["unengaged_unknown_days"]:
            print("WARNING: no messages in metadata fallback — engagement filter "
                  "NOT applied on:", ", ".join(sorted(report["unengaged_unknown_days"])))

    # Keep only the top-N intents by total chat volume across all windows.
    volume = defaultdict(int)
    for bp in args.bp_days:
        for stats in per_bp[bp].values():
            if stats:
                for intent, (total, _esc) in stats.items():
                    # OutOfDomain exists only where native intent classification
                    # ran (Aug 8+), so its early cells are structurally empty.
                    if intent not in ("ALL", "(no intent)", "OutOfDomain"):
                        volume[intent] += total
    top = sorted(volume, key=volume.get, reverse=True)[:args.top_intents]
    print("top intents by volume:",
          ", ".join(f"{i} ({volume[i]})" for i in top))
    intents = {"ALL", *top}

    def order(i):  # ALL first, then by volume
        return (0 if i == "ALL" else 1, -volume.get(i, 0))

    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Escalation rates"
    header = ["BP Day", "Intent"] + [f"T{r:+d}" if r else "T" for r in REL_DAYS]
    ws.append(header)
    for bp in args.bp_days:
        for intent in sorted(intents, key=order):
            row = [bp.isoformat(), intent]
            for rel in REL_DAYS:
                stats = per_bp[bp][rel]
                cell = None
                if stats and intent in stats and stats[intent][0] >= args.min_chats:
                    total, esc = stats[intent]
                    cell = round(esc / total, 4)
                row.append(cell)
            ws.append(row)
            for c in ws[ws.max_row][2:]:
                c.number_format = "0.0%"
    ws.freeze_panes = "C2"

    # Sample sizes behind every rate, for judging significance.
    ws2 = wb.create_sheet("Chat counts")
    ws2.append(header)
    for bp in args.bp_days:
        for intent in sorted(intents, key=order):
            row = [bp.isoformat(), intent]
            for rel in REL_DAYS:
                stats = per_bp[bp][rel]
                row.append(stats[intent][0] if stats and intent in stats else None)
            ws2.append(row)
    ws2.freeze_panes = "C2"

    out.parent.mkdir(exist_ok=True)
    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    main()
