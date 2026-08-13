#!/usr/bin/env python3
"""Bucket exported transcripts and measure where escalations concentrate.

Reads every transcripts_*.xlsx in transcripts/ (Tickets sheet), buckets tickets
by intent x {integration type, autopay status, fulfillment status, service
issue code}, and writes escalation_analysis.json for the heatmap report.

Definitions:
  escalated  - the agent called create_ticket during the conversation
               (the `escalated` column from pull_agent_transcripts.py)
  SOP gap    - escalated AND the escalation looks unguided by a specific rule:
               ticketBodyTag == "general_escalation" (filed under the catch-all),
               OR cf_is_out_of_scope is true, OR intent == "OutOfDomain".
               A proxy, not ground truth: policy-driven escalations carry a
               specific scenario tag; these three signals mark the ones that
               did not match a specific scenario.
"""

import json
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

HERE = Path(__file__).resolve().parent
TRANSCRIPTS_DIR = HERE / "transcripts"
OUT_PATH = HERE / "escalation_analysis.json"
TAXONOMY_PATH = HERE / "tag_taxonomy.json"  # giga tags list -f json > tag_taxonomy.json

# Days excluded from analysis (workbooks stay on disk, they are just not read).
# 2026-08-04: data flagged inaccurate.
EXCLUDE_DAYS = {"2026-08-04"}

ROW_DIM = "intent"
COL_DIMS = {
    "integrationType": {"label": "Integration type", "top": 8},
    "autopayStatus": {"label": "Autopay status", "top": 8},
    "fulfillmentStatus": {"label": "Fulfillment status", "top": 8},
    "serviceIssueCode": {"label": "Service issue code", "top": 8},
}
MIN_ROW_TICKETS = 15  # intents smaller than this fold into "Other"


def norm(value, missing):
    s = "" if value is None else str(value).strip()
    return missing if s in ("", "None") else s


def is_sop_gap(row):
    return (
        norm(row.get("ticketBodyTag"), "") == "general_escalation"
        or str(row.get("cf_is_out_of_scope")) == "True"
        or norm(row.get("intent"), "") == "OutOfDomain"
    )


def load_tickets():
    rows, days = [], []
    for path in sorted(TRANSCRIPTS_DIR.glob("transcripts_*.xlsx")):
        day = path.stem.removeprefix("transcripts_")
        if day in EXCLUDE_DAYS:
            print(f"excluding {day} ({path.name})")
            continue
        sheet = load_workbook(path, read_only=True)["Tickets"]
        it = sheet.iter_rows(values_only=True)
        header = next(it)
        missing = [c for c in ("intent", "escalated", *COL_DIMS) if c not in header]
        if missing:
            print(f"warning: {path.name} lacks {missing} — re-export it "
                  f"(pull_daily_transcripts.py --force) and re-run", file=sys.stderr)
            continue
        rows.extend(dict(zip(header, r)) for r in it)
        days.append(path.stem.removeprefix("transcripts_"))
    if not rows:
        sys.exit(f"error: no usable workbooks in {TRANSCRIPTS_DIR}")
    return rows, days


def bucket_values(rows, dim, top, missing_label):
    counts = Counter(norm(r.get(dim), missing_label) for r in rows)
    keep = [v for v, _ in counts.most_common(top)]
    return keep, counts


def load_tag_taxonomy():
    """tag name -> intent name, from `giga tags list -f json`. Empty if absent."""
    if not TAXONOMY_PATH.exists():
        print(f"note: {TAXONOMY_PATH.name} not found — intents will not be "
              f"derived from tags for pre-analysis-era tickets", file=sys.stderr)
        return {}
    # GreetingOnly / RequestHumanSupport describe the conversation's shape, not
    # its topic (both map to OutOfDomain, which the SOP-gap proxy reads as a
    # gap) — never derive a topical intent from them.
    skip = {"GreetingOnly", "RequestHumanSupport"}
    return {t["name"]: t["intentName"]
            for t in json.loads(TAXONOMY_PATH.read_text())["tags"]
            if t["name"] not in skip}


def main():
    rows, days = load_tickets()

    # Noise filter: drop conversations with at most one message (counting all
    # roles — agent, tool, system included).
    before = len(rows)
    rows = [r for r in rows if int(r.get("messageCount") or 0) > 1]
    print(f"noise filter: removed {before - len(rows)} conversations with <=1 message "
          f"({(before - len(rows)) / before:.1%} of {before})")

    tag_to_intent = load_tag_taxonomy()
    derived = 0
    for r in rows:
        # Intent classification only exists on recent tickets; before that,
        # derive it from the first analysis tag the taxonomy still knows.
        # (Before _gap: the SOP-gap proxy reads the intent.)
        if not norm(r.get("intent"), ""):
            for tag in (norm(r.get("tags"), "") or "").split(", "):
                if tag in tag_to_intent:
                    r["intent"] = tag_to_intent[tag]
                    derived += 1
                    break
        r["_esc"] = str(r.get("escalated")) == "True"
        r["_gap"] = r["_esc"] and is_sop_gap(r)
    if derived:
        print(f"derived intent from tags for {derived} tickets")

    intents, intent_counts = bucket_values(rows, ROW_DIM, 99, "(no intent)")
    row_labels = [v for v in intents if intent_counts[v] >= MIN_ROW_TICKETS]
    if len(row_labels) < len(intents):
        row_labels.append("Other")

    def row_key(r):
        v = norm(r.get(ROW_DIM), "(no intent)")
        return v if v in row_labels else "Other"

    # Order rows hottest-first so the top-left corner is the story.
    esc_by_row = Counter(row_key(r) for r in rows if r["_esc"])
    row_labels.sort(key=lambda v: -esc_by_row[v])

    matrices = {}
    for dim, spec in COL_DIMS.items():
        missing_label = "(none)"
        keep, _ = bucket_values(rows, dim, spec["top"], missing_label)

        def col_key(r):
            v = norm(r.get(dim), missing_label)
            return v if v in keep else "Other"

        col_labels = list(keep)
        if any(col_key(r) == "Other" for r in rows):
            col_labels.append("Other")

        cells = defaultdict(lambda: {"n": 0, "e": 0, "g": 0})
        for r in rows:
            c = cells[(row_key(r), col_key(r))]
            c["n"] += 1
            c["e"] += r["_esc"]
            c["g"] += r["_gap"]
        matrices[dim] = {
            "label": spec["label"],
            "cols": col_labels,
            "cells": [[cells.get((rk, ck), {"n": 0, "e": 0, "g": 0}) for ck in col_labels]
                      for rk in row_labels],
        }

    result = {
        "days": days,
        "totals": {
            "tickets": len(rows),
            "escalated": sum(r["_esc"] for r in rows),
            "sopGap": sum(r["_gap"] for r in rows),
        },
        "rowDim": ROW_DIM,
        "rows": row_labels,
        "matrices": matrices,
    }
    OUT_PATH.write_text(json.dumps(result, indent=1))
    t = result["totals"]
    print(f"{t['tickets']} tickets over {len(days)} day(s): "
          f"{t['escalated']} escalated ({t['escalated']/t['tickets']:.1%}), "
          f"{t['sopGap']} SOP-gap ({t['sopGap']/max(t['escalated'],1):.1%} of escalations)")
    print(f"wrote {OUT_PATH}")


if __name__ == "__main__":
    main()
