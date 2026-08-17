#!/usr/bin/env python3
"""Classify raw transcripts: escalation, cause (LLM), and policy scenario.

For every Customer Support Chat Agent conversation in raw_transcripts/<date>/:

1. escalated?           deterministic — the agent called create_ticket
2. escalation cause     LLM (Claude): "policy_driven" (the policies instructed
                        the escalation) vs "missing_instructions" (no rule or
                        scenario covered the customer's issue)
3. scenario             deterministic — analysis tags / escalation reason
                        mapped to the shared/policies file the conversation
                        belongs to (no LLM, per the tag taxonomy)

Output: one Excel workbook with a date-wise Summary sheet (total chats,
escalated, policy-driven vs missing-instructions), a Scenarios sheet, and a
per-escalation Details sheet.

Requirements:
  - raw transcripts downloaded via pull_raw_transcripts.py
  - a flex-operations clone for the policy texts (--policies-dir)
  - ANTHROPIC_API_KEY set (or `ant auth login`) for the LLM step;
    --dry-run skips the LLM and fills the cause columns as "pending"

LLM results are cached in classification_cache/<date>.jsonl — reruns only
classify tickets not yet cached, so interrupted runs resume for free.

Examples:
  python3 classify_transcripts.py 2026-08-10 2026-08-10 --dry-run
  python3 classify_transcripts.py 2026-08-01 2026-08-11 \
      --policies-dir ~/src/flex-operations --workers 4
"""

import argparse
import json
import re
import sys
import threading
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import date, timedelta
from pathlib import Path

HERE = Path(__file__).resolve().parent
AGENT_NAME = "Customer Support Chat Agent"
AGENT_TEMPLATE_ID = "24bd649e-6a44-4efd-af97-bfaf4dadfb43"
DEFAULT_MODEL = "claude-fable-5"

# ---------------------------------------------------------------------------
# Scenario mapping: analysis tag -> shared/policies git_name.
# Hand-curated from tag_taxonomy.json against the policy frontmatter; edit
# freely — anything unmapped falls back to the escalation-reason map, then
# "(unclassified)".
TAG_TO_POLICY = {
    "AccountIneligibleInfo": "flex-rent-denied-application",
    "AccountPauseAndResume": "account-suspension-reactivation",
    "AccountSuspensionInfo": "account-suspension-reactivation",
    "AccuseFlexFraud": "identity-theft-fraud",
    "AddBackupCard": "switching-payment-methods",
    "AddOrUpdatePaymentMethod": "switching-payment-methods",
    "AddProperty": "moving-authenticated-existing-to-new-property",
    "AddResidentPortal": "property-integration-types",
    "AdjustDownpaymentAmount": "scheduling-rescheduling-payments-chat",
    "AdjustRepaymentAmount": "scheduling-rescheduling-payments-chat",
    "AdverseActionLetterRequest": "credit-lending-regulations",
    "AppDownloadOrInstallIssue": "app-issues-accessibility",
    "AppFreezingIssue": "app-issues-accessibility",
    "ApplicationDenialIdentityVerification": "flex-rent-denied-application",
    "ApplicationDenialInfo": "flex-rent-denied-application",
    "ApplyForFlex": "general-flex-information",
    "CancelAccountAfterFailure": "membership-cancellation-chat",
    "CancelAccountVoluntary": "membership-cancellation-chat",
    "CannotEnterContact": "app-issues-accessibility",
    "CannotLogin": "app-issues-accessibility",
    "CannotReceiveVerificationCode": "email-otp-new-device-login",
    "CannotSignup": "app-issues-accessibility",
    "CardProcessingFeeInfo": "fees-charges-chat",
    "CreditLineDecreaseInfo": "flex-rent-credit-line",
    "CreditLineGeneralInfo": "flex-rent-credit-line",
    "CreditLineIncreaseRequest": "flex-rent-credit-line",
    "CreditReportingInfo": "credit-score-rent-reporting-chat-v2",
    "DebtCollectionsInfo": "credit-reporting-debt-collection",
    "DeleteAccountAndData": "privacy-data-rights",
    "DisputedCharge": "refunds-charged-but-not-paid",
    "DownPaymentInfo": "fees-charges-chat",
    "EvictionConcern": "escalation-handling-chat",
    "FeeFeedback": "fees-charges-chat",
    "FeeInfo": "fees-charges-chat",
    "FlexAnywhereSupport": "using-flex-to-pay-rent-non-integrated-account",
    "FlexBalanceMismatch": "payment-issues-service-issues-chat",
    "HowFlexWorks": "general-flex-information",
    "IncomeVerificationIssue": "flex-rent-denied-application",
    "LanguageSupport": "general-flex-information",
    "LinkCardIssue": "switching-payment-methods",
    "ListedPropertyOrPortalIssue": "property-integration-types",
    "MicrodepositSupport": "property-integration-types",
    "MissedPaymentInfo": "missed-payments-rent-not-paid",
    "MonthlyMembershipFeeInfo": "fees-charges-chat",
    "MoveOutSupport": "moving-authenticated-existing-to-new-property",
    "NotSeeingPropertyOrPortal": "property-integration-types",
    "NsfErrorDespiteFundAvailable": "payment-issues-service-issues-chat",
    "PastDueFlexBalanceInfo": "missed-payments-rent-not-paid",
    "PastDueRentInfo": "missed-payments-rent-not-paid",
    "PayDownPaymentAfterCutoff": "scheduling-rescheduling-payments-chat",
    "PaymentDeclinedInfo": "payment-issues-service-issues-chat",
    "PaymentHardshipOrInsufficientFunds": "financial-hardship-recovery",
    "ProcessDownpaymentManually": "using-flex-to-pay-rent-di-chat",
    "ProcessPayment": "using-flex-to-pay-rent-di-chat",
    "PropertyManagementUpdate": "property-integration-types",
    "ProvideFeedback": "general-flex-information",
    "ReactivateFromBankruptcySuspension": "account-suspension-reactivation",
    "ReactivateFromGeneralSuspension": "account-suspension-reactivation",
    "ReactivateVoluntaryReturn": "account-suspension-reactivation",
    "ReactivationDenialInfo": "account-suspension-reactivation",
    "RefundDownpayment": "refunds-charged-but-not-paid",
    "RefundFee": "refunds-charged-but-not-paid",
    "RefundRepayment": "refunds-charged-but-not-paid",
    "RemovePaymentMethod": "switching-payment-methods",
    "RemoveUtilityFeeFromRentAmount": "using-flex-to-pay-rent-di-chat",
    "RentConfirmedInAppNotPortal": "property-integration-types",
    "RentNotInitiatedAfterDownpayment": "payment-issues-service-issues-chat",
    "RentNotInitiatedBeforeDownpayment": "payment-issues-service-issues-chat",
    "RentPaymentStatus": "using-flex-to-pay-rent-di-chat",
    "RentStatementRequest": "using-flex-to-pay-rent-di-chat",
    "RentTransactionRejectedInfo": "payment-issues-service-issues-chat",
    "RenterPropertyUpdate": "updating-customer-account-information-chat",
    "RepaymentInfo": "scheduling-rescheduling-payments-chat",
    "ReportFraudConcern": "identity-theft-fraud",
    "ReportIdentityTheft": "identity-theft-fraud",
    "RequestHumanSupport": "escalation-handling-chat",
    "RescheduleDownpayment": "scheduling-rescheduling-payments-chat",
    "RescheduleRepayment": "scheduling-rescheduling-payments-chat",
    "ResetPassword": "app-issues-accessibility",
    "RetrieveUsername": "app-issues-accessibility",
    "SelfSubmitSupport": "using-flex-to-pay-rent-non-integrated-account",
    "SettleOutstandingFlexBalance": "missed-payments-rent-not-paid",
    "SplitAcrossMultiplePaymentMethods": "switching-payment-methods",
    "SplitIntoMoreInstallments": "scheduling-rescheduling-payments-chat",
    "SplitWithOthersInfo": "general-flex-information",
    "StopOrPauseAutopay": "rent-autopay-management-chat",
    "SwitchFundingSource": "switching-payment-methods",
    "UnsubscribeOrDnc": "account-settings-communication-preferences",
    "UpdateProfile": "updating-customer-account-information-chat",
    "UpdateRentAmount": "using-flex-to-pay-rent-di-chat",
    "VirtualAccountIssue": "using-flex-to-pay-rent-non-integrated-account",
    "WithdrawFlexFunds": "flex-funds",
    # Shape-of-conversation tags carry no topical scenario:
    "GreetingOnly": None,
}
# Escalation reason (create_ticket ticket_body_tag) -> policy, used when no
# analysis tag mapped.
REASON_TO_POLICY = {
    "payment_issue": "payment-issues-service-issues-chat",
    "payment_method_issue": "switching-payment-methods",
    "eligibility_override": "lower-first-payment-offer",
    "general_escalation": "escalation-handling-chat",
    "account_update_request": "updating-customer-account-information-chat",
    "refund_request": "refunds-charged-but-not-paid",
    "reschedule_payment": "scheduling-rescheduling-payments-chat",
    "app_issue": "app-issues-accessibility",
    "cc_control_center": "control-center-flex-for-everyday-bills-v2",
    "hardship_request": "financial-hardship-recovery",
    "rent_amount_issue": "using-flex-to-pay-rent-di-chat",
    "missed_payment": "missed-payments-rent-not-paid",
    "cbnp_escalation": "refunds-charged-but-not-paid",
    "credit_line_adjustment": "flex-rent-credit-line",
    "fee_inquiry": "fees-charges-chat",
    "frustrated_customer": "escalation-handling-chat",
    "eviction_related_language": "escalation-handling-chat",
    "fraud_escalation": "identity-theft-fraud",
    "repeated_agent_request": "escalation-handling-chat",
    "emergency": "escalation-handling-chat",
    "regulatory_dispute": "credit-lending-regulations",
    "legal_representation": "legal-representation-agreements",
    "active_duty_request": "active-duty-procedures-scra-mla",
    "direct_deposit_inquiry": "updating-customer-account-information-chat",
    "request_to_pay_membership_fee": "fees-charges-chat",
}

ROLE = {1: "CUSTOMER", 2: "AGENT", 3: "SYSTEM", 4: "TOOL", 5: "HUMAN_AGENT"}

CAUSE_SCHEMA = {
    "type": "object",
    "properties": {
        "cause": {"type": "string", "enum": ["policy_driven", "missing_instructions"]},
        "confidence": {"type": "string", "enum": ["high", "medium", "low"]},
        "reason": {"type": "string",
                   "description": "One or two sentences citing the rule/scenario that "
                                  "mandated the escalation, or the gap that forced it."},
    },
    "required": ["cause", "confidence", "reason"],
    "additionalProperties": False,
}


def load_tickets(raw_dir: Path, day: str):
    for p in sorted((raw_dir / day).glob("*.json")):
        try:
            t = json.loads(p.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        if t.get("agentTemplateId") == AGENT_TEMPLATE_ID or t.get("agentName") == AGENT_NAME:
            yield t


def escalation_info(ticket):
    """(escalated, ticket_body_tag) from create_ticket tool calls."""
    for m in ticket.get("messages") or []:
        for tc in m.get("toolCalls") or []:
            fn = tc.get("function") or {}
            if (tc.get("name") or fn.get("name")) == "create_ticket":
                try:
                    return True, json.loads(fn.get("arguments") or "{}").get("ticket_body_tag", "")
                except json.JSONDecodeError:
                    return True, ""
    return False, ""


def classify_scenario(ticket, reason_tag):
    tags = (ticket.get("analysis") or {}).get("tags") or []
    # Topical tags first; RequestHumanSupport only describes the conversation's shape.
    for tag in sorted(tags, key=lambda t: t == "RequestHumanSupport"):
        if TAG_TO_POLICY.get(tag):
            return TAG_TO_POLICY[tag]
    if reason_tag and reason_tag in REASON_TO_POLICY:
        return REASON_TO_POLICY[reason_tag]
    return "(unclassified)"


def render_transcript(ticket, max_chars=12_000):
    lines = []
    for m in ticket.get("messages") or []:
        role = ROLE.get(m.get("role"), "?")
        content = m.get("content")
        if role == "TOOL":
            continue  # tool result payloads are noise for cause classification
        if isinstance(content, str) and content.strip():
            lines.append(f"{role}: {content.strip()[:1500]}")
        for tc in m.get("toolCalls") or []:
            fn = tc.get("function") or {}
            name = tc.get("name") or fn.get("name")
            if name:
                args = (fn.get("arguments") or "")[:400]
                lines.append(f"AGENT->TOOL: {name}({args})")
    text = "\n".join(lines)
    return text[:max_chars]


class CauseClassifier:
    """LLM classification of escalation cause, with cached policy context."""

    def __init__(self, policies_dir: Path, model: str):
        import anthropic  # deferred so --dry-run works without the SDK/key
        self.client = anthropic.Anthropic()
        self.model = model
        self.policies_dir = policies_dir
        agent_pol = policies_dir / "agents/customer-support-chat-agent/policies"
        rules = (agent_pol / "rules.md").read_text(encoding="utf-8")
        esc = (policies_dir / "shared/policies/escalation-handling-chat.md").read_text(encoding="utf-8")
        index = []
        for f in sorted((policies_dir / "shared/policies").glob("*.md")):
            m = re.search(r"^git_name:\s*(\S+).*?^description:\s*(.+?)^---",
                          f.read_text(encoding="utf-8"), re.S | re.M)
            if m:
                desc = " ".join(m.group(2).split())[:300]
                index.append(f"- {m.group(1)}: {desc}")
        self.system = [{
            "type": "text",
            "text": (
                "You audit escalations made by Flex's Customer Support Chat Agent. "
                "The agent escalates by calling create_ticket. Classify WHY this "
                "conversation was escalated:\n\n"
                "- policy_driven: the policies below instructed the escalation for this "
                "situation — an Auto-Escalation topic, a scenario step that says to create "
                "a ticket, or the Escalation Handling SOP's path for a customer who insists "
                "after the agent attempted to help.\n"
                "- missing_instructions: no rule or scenario covered the customer's actual "
                "issue — the agent escalated because the policies gave it no way to resolve "
                "or even address the request (missing scenario, missing information, or a "
                "situation the SOPs never anticipated).\n\n"
                "If the agent escalated in violation of the policies (e.g. without a "
                "required resolution attempt) but a policy DID cover the issue, that is "
                "still policy_driven territory — missing_instructions is strictly about "
                "gaps in coverage.\n\n"
                "=== AGENT RULES (rules.md) ===\n" + rules +
                "\n\n=== ESCALATION HANDLING SOP ===\n" + esc +
                "\n\n=== INDEX OF ALL SHARED POLICY SCENARIOS ===\n" + "\n".join(index)
            ),
            "cache_control": {"type": "ephemeral"},
        }]
        self._policy_cache = {}

    def policy_text(self, slug):
        if slug not in self._policy_cache:
            f = self.policies_dir / "shared/policies" / f"{slug}.md"
            self._policy_cache[slug] = f.read_text(encoding="utf-8")[:8000] if f.exists() else ""
        return self._policy_cache[slug]

    def classify(self, ticket, scenario, reason_tag):
        parts = []
        if scenario not in ("(unclassified)", "escalation-handling-chat") and self.policy_text(scenario):
            parts.append(f"=== MATCHED SCENARIO POLICY ({scenario}) ===\n{self.policy_text(scenario)}\n")
        parts.append(f"Escalation reason tag the agent filed: {reason_tag or '(none)'}\n")
        parts.append("=== TRANSCRIPT ===\n" + render_transcript(ticket))
        response = self.client.messages.create(
            model=self.model,
            max_tokens=8000,  # thinking counts toward max_tokens on this model
            output_config={"effort": "low",
                           "format": {"type": "json_schema", "schema": CAUSE_SCHEMA}},
            system=self.system,
            messages=[{"role": "user", "content": "\n".join(parts)}],
        )
        if response.stop_reason == "refusal":
            return {"cause": "llm_refused", "confidence": "low", "reason": "model refused"}
        text = next(b.text for b in response.content if b.type == "text")
        return json.loads(text)


def main():
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("start", type=date.fromisoformat)
    parser.add_argument("end", type=date.fromisoformat)
    parser.add_argument("--raw-dir", type=Path, default=HERE / "raw_transcripts")
    parser.add_argument("--policies-dir", type=Path, default=HERE / "flex-operations",
                        help="flex-operations clone (for policy texts)")
    parser.add_argument("--out", type=Path, default=None,
                        help="Output workbook (default results/escalation_classification_<range>.xlsx)")
    parser.add_argument("--model", default=DEFAULT_MODEL)
    parser.add_argument("--workers", type=int, default=4)
    parser.add_argument("--limit", type=int, default=None,
                        help="Max LLM calls this run (trial/cost control); rest stay pending")
    parser.add_argument("--cache-dir", type=Path, default=HERE / "classification_cache")
    parser.add_argument("--dry-run", action="store_true",
                        help="Skip the LLM; cause columns become 'pending'")
    args = parser.parse_args()

    out = args.out or HERE / "results" / \
        f"escalation_classification_{args.start}_{args.end}.xlsx"
    args.cache_dir.mkdir(exist_ok=True)

    classifier = None
    if not args.dry_run:
        classifier = CauseClassifier(args.policies_dir, args.model)

    day_rows = []      # per-day summary
    detail_rows = []   # per-escalated-ticket detail
    scen_rows = Counter()  # (date, scenario) -> [chats, escalated]
    llm_budget = [args.limit if args.limit is not None else float("inf")]
    lock = threading.Lock()

    day = args.start
    while day <= args.end:
        d = day.isoformat()
        if not (args.raw_dir / d).exists():
            print(f"[{d}] no raw transcripts — skipping", file=sys.stderr)
            day += timedelta(days=1)
            continue

        cache_file = args.cache_dir / f"{d}.jsonl"
        cached = {}
        if cache_file.exists():
            for line in cache_file.read_text().splitlines():
                try:
                    rec = json.loads(line)
                    cached[rec["ticketId"]] = rec
                except json.JSONDecodeError:
                    pass

        total = 0
        escalated_tickets = []
        for t in load_tickets(args.raw_dir, d):
            total += 1
            esc, reason_tag = escalation_info(t)
            scenario = classify_scenario(t, reason_tag)
            key = (d, scenario)
            scen_rows[key] = scen_rows.get(key, (0, 0))
            scen_rows[key] = (scen_rows[key][0] + 1, scen_rows[key][1] + int(esc))
            if esc:
                escalated_tickets.append((t, scenario, reason_tag))

        def run_one(item):
            t, scenario, reason_tag = item
            tid = t.get("id", "")
            if tid in cached:
                return {**cached[tid], "scenario": scenario, "reasonTag": reason_tag}
            with lock:
                if llm_budget[0] <= 0 or classifier is None:
                    return {"ticketId": tid, "cause": "pending", "confidence": "",
                            "reason": "", "scenario": scenario, "reasonTag": reason_tag}
                llm_budget[0] -= 1
            try:
                result = classifier.classify(t, scenario, reason_tag)
            except Exception as exc:  # noqa: BLE001 — record and move on
                return {"ticketId": tid, "cause": "llm_error", "confidence": "low",
                        "reason": str(exc)[:200], "scenario": scenario, "reasonTag": reason_tag}
            rec = {"ticketId": tid, **result}
            with lock:
                with cache_file.open("a") as fh:
                    fh.write(json.dumps(rec) + "\n")
            return {**rec, "scenario": scenario, "reasonTag": reason_tag}

        results = []
        with ThreadPoolExecutor(max_workers=args.workers) as pool:
            for fut in as_completed([pool.submit(run_one, item) for item in escalated_tickets]):
                results.append(fut.result())

        counts = Counter(r["cause"] for r in results)
        day_rows.append({
            "date": d, "total_chats": total, "escalated": len(escalated_tickets),
            "policy_driven": counts.get("policy_driven", 0),
            "missing_instructions": counts.get("missing_instructions", 0),
            "pending_or_error": counts.get("pending", 0) + counts.get("llm_error", 0)
            + counts.get("llm_refused", 0),
        })
        for r in results:
            detail_rows.append({"date": d, **r})
        print(f"[{d}] chats={total} escalated={len(escalated_tickets)} "
              f"policy={counts.get('policy_driven', 0)} "
              f"missing={counts.get('missing_instructions', 0)} "
              f"pending/err={day_rows[-1]['pending_or_error']}", flush=True)
        day += timedelta(days=1)

    # ---- Excel ----
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["date", "total_chats", "escalated", "escalation_rate",
               "escalated_policy_driven", "escalated_missing_instructions",
               "pending_or_error"])
    for r in day_rows:
        rate = r["escalated"] / r["total_chats"] if r["total_chats"] else 0
        ws.append([r["date"], r["total_chats"], r["escalated"], round(rate, 4),
                   r["policy_driven"], r["missing_instructions"], r["pending_or_error"]])
    ws.freeze_panes = "A2"

    # One aggregate row per scenario across the whole range.
    scen_totals = {}
    for (_d, scenario), (chats, esc) in scen_rows.items():
        cur = scen_totals.setdefault(scenario, [0, 0])
        cur[0] += chats
        cur[1] += esc
    cause_by_scen = Counter()
    for r in detail_rows:
        cause_by_scen[(r["scenario"], r["cause"])] += 1

    ws2 = wb.create_sheet("Scenarios")
    ws2.append(["scenario_policy", "total_chats", "total_escalations",
                "escalations_agent_instructed", "escalations_uninformed",
                "pending_or_error"])
    for scenario, (chats, esc) in sorted(scen_totals.items(), key=lambda kv: -kv[1][1]):
        instructed = cause_by_scen.get((scenario, "policy_driven"), 0)
        uninformed = cause_by_scen.get((scenario, "missing_instructions"), 0)
        ws2.append([scenario, chats, esc, instructed, uninformed,
                    esc - instructed - uninformed])
    ws2.freeze_panes = "A2"

    ws3 = wb.create_sheet("Details")
    ws3.append(["date", "ticketId", "scenario_policy", "escalation_reason_tag",
                "cause", "confidence", "reason"])
    for r in detail_rows:
        ws3.append([r["date"], r["ticketId"], r["scenario"], r.get("reasonTag", ""),
                    r["cause"], r.get("confidence", ""), str(r.get("reason", ""))[:500]])
    ws3.freeze_panes = "A2"

    out.parent.mkdir(exist_ok=True)
    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    main()
